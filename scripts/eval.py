#!/usr/bin/env python3
"""
Excel AI — Comprehensive Model Evaluation

Tests Ollama models on real-world Excel use cases extracted from the demo spreadsheet.
Evaluates accuracy, quality, and latency across 10 task categories.

Usage:
    python3 scripts/eval.py                          # Run full evaluation
    python3 scripts/eval.py --models llama3.2:1b gemma3:4b   # Test specific models
    python3 scripts/eval.py --cleanup                # Remove each model after testing
    python3 scripts/eval.py --output results.xlsx    # Custom output path
"""

import argparse
import json
import os
import re
import subprocess
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

# ── Constants ──────────────────────────────────────────────────────────────

OLLAMA_URL = os.environ.get("OLLAMA_URL", "http://127.0.0.1:11434")
SYSTEM_PROMPT = (
    "You are a helpful assistant embedded in Excel. "
    "Give concise answers suitable for spreadsheet cells."
)
TEMPERATURE = 0.3
MAX_TOKENS = 512
NUM_CTX = 2048
TIMEOUT = 90

ALL_MODELS = [
    # Tier 1: Ultra-small (~1B)
    "llama3.2:1b",
    "gemma3:1b",
    "qwen2.5:1.5b",
    "qwen3:0.6b",
    "qwen3:1.7b",
    "smollm2:1.7b",
    # Tier 2: Small (~3-4B)
    "llama3.2:3b",
    "gemma3:4b",
    "phi4-mini",
    "qwen2.5:3b",
    "qwen3:4b",
    "granite3-moe:3b",
    # Tier 3: New generation
    "gemma4:e2b",
    "gemma4:e4b",
]


# ── Test Cases (from excel_ai_demo.xlsx) ───────────────────────────────────

SENTIMENT_TESTS = [
    ("Absolutely love this product! Best purchase I've made all year.", "Positive"),
    ("Terrible quality. Broke after two days. Want a refund immediately.", "Negative"),
    ("It's okay, nothing special. Does what it says.", "Neutral"),
    ("Customer support was incredibly helpful and resolved my issue fast!", "Positive"),
    ("Shipping took 3 weeks. The product itself is fine I guess.", "Neutral"),
    ("DO NOT BUY. Scam product, nothing like the photos.", "Negative"),
    ("Exceeded my expectations. Already ordered two more for gifts.", "Positive"),
    ("It works but the instructions were confusing.", "Neutral"),
    ("I've been using this for 6 months now, still going strong.", "Positive"),
    ("Wrong item was sent. Very frustrating experience.", "Negative"),
    ("Meh. Wouldn't buy again, but I won't return it either.", "Neutral"),
    ("Five stars! This changed my morning routine completely.", "Positive"),
]

EXTRACTION_COMPANY_TESTS = [
    ("Hi, this is John from Acme Corp. We'd like to place a $4,500 order by March 15th.", "Acme Corp"),
    ("Following up on our call — Tesla confirmed a $2.3M investment round closing Q2 2025.", "Tesla"),
    ("Invoice #9921 from Bright Solutions LLC, total due: $12,340.00, payment by 04/01/2025.", "Bright Solutions LLC"),
    ("The board at Nextera Energy approved a $850K budget for the new facility on Jan 8.", "Nextera Energy"),
    ("Per our agreement, Globex Inc. owes $67,000 no later than June 30th, 2025.", "Globex Inc"),
    ("Amazon reported $143.3 billion in net sales for Q4 2024, beating analyst expectations.", "Amazon"),
    ("Please wire $5,250 to Pinnacle Systems by end of business Friday, Feb 28, 2025.", "Pinnacle Systems"),
    ("Memo: Initech has allocated $320K toward R&D for the fiscal year starting April 1.", "Initech"),
]

EXTRACTION_AMOUNT_TESTS = [
    ("Hi, this is John from Acme Corp. We'd like to place a $4,500 order by March 15th.", "4500"),
    ("Following up on our call — Tesla confirmed a $2.3M investment round closing Q2 2025.", "2300000"),
    ("Invoice #9921 from Bright Solutions LLC, total due: $12,340.00, payment by 04/01/2025.", "12340"),
    ("The board at Nextera Energy approved a $850K budget for the new facility on Jan 8.", "850000"),
    ("Per our agreement, Globex Inc. owes $67,000 no later than June 30th, 2025.", "67000"),
    ("Amazon reported $143.3 billion in net sales for Q4 2024, beating analyst expectations.", "143300000000"),
    ("Please wire $5,250 to Pinnacle Systems by end of business Friday, Feb 28, 2025.", "5250"),
    ("Memo: Initech has allocated $320K toward R&D for the fiscal year starting April 1.", "320000"),
]

CATEGORIZATION_TESTS = [
    ("Uber ride to SFO airport", "Travel"),
    ("AWS monthly hosting bill", "Software"),
    ("Team lunch at Olive Garden (8 people)", "Meals"),
    ("Adobe Creative Cloud annual subscription", "Software"),
    ("Office chair — ergonomic, for home office", "Office"),
    ("Delta Airlines LAX→JFK roundtrip", "Travel"),
    ("Mailchimp email marketing — Pro plan", "Marketing"),
    ("WeWork hot desk — January", "Office"),
    ("Client gift basket — holiday thank you", "Marketing"),
    ("Zoom Pro yearly subscription", "Software"),
    ("Parking garage — downtown client meeting", "Travel"),
    ("Google Workspace Business Standard", "Software"),
    ("Professional headshots for LinkedIn", "Marketing"),
    ("Office supplies — printer paper, toner, pens", "Office"),
]

CLEANING_TESTS = [
    ("john.doe AT gmail DOT com", "john.doe@gmail.com"),
    ("(555) 123-4567 ext. 89", "+15551234567"),
    ("New  York ,  NY   10001", "New York, NY 10001"),
    ("MICROSOFT CORPORATION INC.", "Microsoft"),
    ("2025/15/03", "03/15/2025"),
    ("$1,234,567.89 USD", "1234567.89"),
    ("Jean-François Müller", "Jean-Francois Muller"),
    ("123 Main St., Ste. 4B, Bldg. A", "123 Main Street, Suite 4B, Building A"),
    ("approx 3.5 kg (7.7 lbs)", "3.5"),
    ("United States of America", "US"),
]

FORMULA_TESTS = [
    ("Look up the price from column B where the product name in column A matches cell D1",
     ["VLOOKUP", "INDEX", "MATCH", "XLOOKUP"]),
    ("Count how many cells in A1:A100 contain the word 'urgent'",
     ["COUNTIF"]),
    ("Calculate the percentage change between cell B2 and B3",
     ["B3", "B2"]),
    ("Find the second largest value in the range C1:C50",
     ["LARGE"]),
    ("Combine first name in A2 and last name in B2 with a space between them",
     ["CONCAT", "&", "TEXTJOIN"]),
    ("Extract just the domain name from an email address in cell A1",
     ["MID", "RIGHT", "FIND", "@"]),
    ("If the value in A1 is over 100, show 'High', between 50-100 show 'Medium', otherwise 'Low'",
     ["IF"]),
    ("Calculate the number of working days between two dates in A1 and B1, excluding holidays in D1:D10",
     ["NETWORKDAYS"]),
    ("Sum all values in column B where column A says 'Marketing'",
     ["SUMIF"]),
    ("Convert a Unix timestamp in A1 to a readable date",
     ["DATE", "/"]),
]

TRANSLATION_TESTS = [
    ("Wireless noise-cancelling headphones with 30-hour battery life.",
     ["inalámbric", "auricular", "batería", "30"]),
    ("Organic green tea, hand-picked from the highlands of Kyoto.",
     ["orgánic", "verde", "Kyoto", "Kioto"]),
    ("Adjustable standing desk, fits monitors up to 32 inches.",
     ["escritorio", "ajustable", "monitor", "32"]),
    ("Stainless steel water bottle — keeps drinks cold for 24 hours.",
     ["acero", "inoxidable", "botella", "24"]),
    ("Professional chef's knife, forged from Japanese carbon steel.",
     ["cuchillo", "chef", "profesional", "japonés", "japon"]),
    ("Compact travel backpack with anti-theft hidden zipper.",
     ["mochila", "viaje", "compact", "cremallera", "antirrobo"]),
    ("Smart LED bulb, voice-controlled, 16 million color options.",
     ["LED", "inteligente", "voz", "16", "color"]),
    ("Warning: contains nuts. May cause allergic reaction.",
     ["advertencia", "nuec", "alérgi", "reacción"]),
]

CODE_TESTS = [
    ("Validate an email address", "Python",
     ["def", "import", "re", "email", "@"]),
    ("FizzBuzz from 1 to 100", "JavaScript",
     ["for", "Fizz", "Buzz", "console", "log", "%"]),
    ("Connect to a PostgreSQL database", "Python",
     ["import", "connect", "psycopg", "host", "database"]),
    ("HTTP GET request and parse JSON response", "curl",
     ["curl", "http", "json", "jq"]),
    ("Sort an array of objects by a 'date' property", "TypeScript",
     ["sort", "date", "=>", "function"]),
    ("Read a CSV file and print the first 5 rows", "Python",
     ["import", "csv", "open", "read", "print"]),
    ("Cron job expression that runs every weekday at 9am EST", "cron",
     ["9", "*", "1-5", "0"]),
    ("Regex to match US phone numbers in any format", "regex",
     ["\\d", "\\(?", "phone", "\\b", "?"]),
]

SUMMARIZATION_TESTS = [
    (
        "Discussed Q1 results which were 12% above forecast. Marketing team proposed "
        "increasing social media budget by $50K for Q2. Engineering flagged that the new "
        "API will be delayed 2 weeks due to a dependency on the auth service rewrite. HR "
        "mentioned three new hires starting in April. CEO wants a revised roadmap by Friday.",
        ["Q1", "12%", "marketing", "API", "delay", "roadmap"],
    ),
    (
        "Client call with Acme Corp. They love the dashboard but want custom exports in CSV "
        "and PDF. Also asked about SSO integration — said it's a dealbreaker for enterprise "
        "tier. Pricing discussion: they pushed back on the $15/seat model, suggested volume "
        "discount above 500 seats. Follow-up meeting scheduled for next Thursday.",
        ["Acme", "CSV", "PDF", "SSO", "pricing", "volume", "discount"],
    ),
    (
        "Sprint retrospective: team velocity improved 20% over last sprint. Two bugs in "
        "production were caused by missing input validation. Agreed to add mandatory code "
        "review for all PRs. DevOps will migrate CI/CD to GitHub Actions by end of month. "
        "Morale is high after the product launch.",
        ["velocity", "20%", "bug", "code review", "CI/CD", "GitHub Actions"],
    ),
    (
        "Board meeting summary: annual revenue hit $4.2M, up 35% YoY. Burn rate is $280K/month "
        "with 18 months runway. Board approved Series A fundraising — targeting $10M at $40M "
        "pre-money. Legal flagged potential IP issue with competitor patent. Next board meeting May 15.",
        ["revenue", "$4.2M", "35%", "Series A", "IP", "patent"],
    ),
    (
        "Customer support weekly review: 847 tickets handled, avg resolution time 4.2 hours "
        "(down from 5.1). Top complaint: password reset flow is confusing — 23% of tickets. "
        "Team suggests adding a self-service reset option. CSAT score 4.1/5. New chatbot "
        "deflected 15% of tier-1 queries.",
        ["847", "resolution", "password", "23%", "chatbot", "CSAT"],
    ),
]

EMAIL_TESTS = [
    ("Sarah Chen", "Stripe", "Met at SaaStr conference, interested in our analytics tool"),
    ("Marcus Johnson", "Shopify", "Signed up for free trial 2 weeks ago, hasn't activated"),
    ("Priya Patel", "HubSpot", "Current customer, contract renewal in 30 days"),
    ("James O'Brien", "Snowflake", "Downloaded our whitepaper on data pipelines"),
    ("Lisa Yamamoto", "Figma", "Referred by mutual connection David Kim"),
    ("Carlos Rivera", "Notion", "Attended our webinar on productivity tools last week"),
]

PRODUCT_TESTS = [
    ("ThermoFlask Pro",
     "64oz, vacuum insulated, stainless steel, dishwasher safe, keeps cold 48hrs",
     ["ThermoFlask", "vacuum", "stainless", "48", "cold"]),
    ("LumiDesk LED Panel",
     "24x18 inches, 5000K daylight, dimmable, USB-C powered, ultra-thin",
     ["LumiDesk", "LED", "5000K", "USB-C", "dimmable"]),
    ("CloudStep Running Shoe",
     "carbon fiber plate, 8.2oz, 35mm stack height, recycled mesh upper",
     ["CloudStep", "carbon", "fiber", "lightweight", "recycled"]),
    ("ZenPod White Noise Machine",
     "30 sounds, timer function, portable, USB rechargeable, < 40dB",
     ["ZenPod", "30", "sound", "timer", "portable"]),
    ("PixelPro Webcam 4K",
     "4K/30fps, auto-focus, built-in ring light, privacy shutter, USB-A",
     ["PixelPro", "4K", "auto-focus", "ring light", "privacy"]),
    ("GreenBlend Smoothie Maker",
     "1200W, 6 stainless blades, BPA-free, self-cleaning mode, 64oz pitcher",
     ["GreenBlend", "1200W", "BPA", "self-cleaning", "blend"]),
]


# ── API Functions ──────────────────────────────────────────────────────────

def call_model(model: str, prompt: str) -> tuple[str, float]:
    """Send a prompt to Ollama and return (response_text, latency_seconds)."""
    body = json.dumps({
        "model": model,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ],
        "temperature": TEMPERATURE,
        "max_tokens": MAX_TOKENS,
        "num_ctx": NUM_CTX,
        "stream": False,
    }).encode()

    req = urllib.request.Request(
        f"{OLLAMA_URL}/v1/chat/completions",
        data=body,
        headers={"Content-Type": "application/json"},
    )

    start = time.monotonic()
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
            data = json.loads(resp.read())
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as e:
        return f"ERROR: {e}", time.monotonic() - start

    elapsed = time.monotonic() - start
    content = ""
    try:
        content = data["choices"][0]["message"]["content"].strip()
    except (KeyError, IndexError):
        content = "ERROR: empty response"
    return content, elapsed


def pull_model(model: str) -> bool:
    """Pull a model via ollama CLI. Returns True on success."""
    print(f"  Pulling {model}...")
    try:
        result = subprocess.run(
            ["ollama", "pull", model],
            capture_output=True, text=True, timeout=1800,
        )
        if result.returncode != 0:
            print(f"  FAILED to pull {model}: {result.stderr.strip()}")
            return False
    except subprocess.TimeoutExpired:
        print(f"  TIMEOUT pulling {model} (30min limit)")
        return False
    except Exception as e:
        print(f"  ERROR pulling {model}: {e}")
        return False
    print(f"  Pulled {model} successfully.")
    return True


def remove_model(model: str):
    """Remove a model to free disk space."""
    subprocess.run(["ollama", "rm", model], capture_output=True, timeout=30)
    print(f"  Removed {model}.")


def get_model_size(model: str) -> str:
    """Get model file size from ollama list."""
    result = subprocess.run(
        ["ollama", "list"], capture_output=True, text=True, timeout=10,
    )
    for line in result.stdout.splitlines():
        if model in line:
            parts = line.split()
            for i, p in enumerate(parts):
                if p in ("MB", "GB", "KB"):
                    return f"{parts[i-1]} {p}"
                if re.match(r"\d+(\.\d+)?\s*(MB|GB|KB)", " ".join(parts[i:i+2])):
                    return " ".join(parts[i:i+2])
    return "?"


# ── Scoring Functions ──────────────────────────────────────────────────────

def score_exact(response: str, expected: str) -> float:
    """Score 1.0 if expected label appears in response (case-insensitive)."""
    resp_lower = response.lower().strip()
    exp_lower = expected.lower().strip()
    # Direct match or starts with expected
    if resp_lower == exp_lower or resp_lower.startswith(exp_lower):
        return 1.0
    # Expected appears as a standalone word in response
    if re.search(r'\b' + re.escape(exp_lower) + r'\b', resp_lower):
        return 1.0
    return 0.0


def score_contains(response: str, expected: str) -> float:
    """Score 1.0 if expected string appears in response (case-insensitive)."""
    if expected.lower() in response.lower():
        return 1.0
    return 0.0


def score_amount(response: str, expected: str) -> float:
    """Score extraction of dollar amounts."""
    resp_clean = response.replace(",", "").replace("$", "").replace(" ", "")
    numbers = re.findall(r'[\d.]+', resp_clean)
    exp_val = float(expected)
    for num_str in numbers:
        try:
            val = float(num_str)
            if abs(val - exp_val) < 1.0:
                return 1.0
            if exp_val != 0 and abs(val - exp_val) / exp_val < 0.01:
                return 1.0
        except ValueError:
            continue
    return 0.0


def score_cleaning(response: str, expected: str) -> float:
    """Score data cleaning accuracy."""
    resp = response.strip().strip('"').strip("'").strip("`")
    exp = expected.strip()
    if resp == exp:
        return 1.0
    if resp.lower() == exp.lower():
        return 0.9
    # Partial credit for close matches
    resp_norm = re.sub(r'\s+', ' ', resp.lower())
    exp_norm = re.sub(r'\s+', ' ', exp.lower())
    if resp_norm == exp_norm:
        return 0.8
    # Check if the core content is present
    if exp_norm in resp_norm:
        return 0.6
    return 0.0


def score_keywords(response: str, keywords: list[str], min_match: int = None) -> float:
    """Score based on fraction of keywords found in response."""
    if not keywords:
        return 0.0
    found = 0
    resp_lower = response.lower()
    for kw in keywords:
        if kw.lower() in resp_lower:
            found += 1
    return found / len(keywords)


def score_formula(response: str, expected_functions: list[str]) -> float:
    """Score Excel formula generation."""
    resp = response.strip()
    score = 0.0
    # Check for formula marker
    if "=" in resp:
        score += 0.3
    # Check for expected functions/references
    found = 0
    for func in expected_functions:
        if func.upper() in resp.upper() or func in resp:
            found += 1
    if expected_functions:
        score += 0.7 * (found / len(expected_functions))
    return min(score, 1.0)


def score_code(response: str, language: str, keywords: list[str]) -> float:
    """Score code snippet generation."""
    resp = response.strip()
    score = 0.0
    # Check for code block or language-specific patterns
    if "```" in resp or any(kw in resp for kw in ["def ", "function ", "import ", "const ", "let "]):
        score += 0.3
    # Check keywords
    found = sum(1 for kw in keywords if kw.lower() in resp.lower())
    if keywords:
        score += 0.7 * (found / len(keywords))
    return min(score, 1.0)


def score_email(response: str, name: str, company: str) -> float:
    """Score email draft quality."""
    resp_lower = response.lower()
    score = 0.0
    first_name = name.split()[0]
    # Mentions contact name
    if first_name.lower() in resp_lower:
        score += 0.35
    # Mentions company
    if company.lower() in resp_lower:
        score += 0.25
    # Has greeting
    if any(g in resp_lower for g in ["hi ", "hello ", "dear ", "hey "]):
        score += 0.15
    # Has professional structure (subject line or sign-off)
    if any(s in resp_lower for s in ["subject:", "best", "regards", "sincerely", "cheers"]):
        score += 0.1
    # Reasonable length (at least 50 chars, not more than 2000)
    if 50 <= len(response) <= 2000:
        score += 0.15
    return min(score, 1.0)


def score_product(response: str, product_name: str, keywords: list[str]) -> float:
    """Score product copy quality."""
    resp_lower = response.lower()
    score = 0.0
    # Mentions product name
    if product_name.lower() in resp_lower:
        score += 0.3
    # Keyword coverage
    found = sum(1 for kw in keywords if kw.lower() in resp_lower)
    if keywords:
        score += 0.4 * (found / len(keywords))
    # Reasonable length
    if 80 <= len(response) <= 1500:
        score += 0.15
    # No excessive formatting artifacts
    if response.count("**") <= 6 and response.count("```") == 0:
        score += 0.15
    return min(score, 1.0)


def score_summary(response: str, keywords: list[str], original_length: int) -> float:
    """Score text summarization quality."""
    score = 0.0
    # Compression: response should be shorter than original
    if len(response) < original_length:
        score += 0.3
    elif len(response) < original_length * 1.5:
        score += 0.15
    # Keyword coverage
    found = sum(1 for kw in keywords if kw.lower() in response.lower())
    if keywords:
        score += 0.5 * (found / len(keywords))
    # Concise (under 300 chars is great for a one-liner)
    if len(response) < 300:
        score += 0.2
    elif len(response) < 500:
        score += 0.1
    return min(score, 1.0)


# ── Test Runners ───────────────────────────────────────────────────────────

def run_sentiment(model: str) -> list[dict]:
    prompt_template = (
        "Classify this customer feedback as Positive, Negative, or Neutral. "
        "Reply with one word only.\n\nContext:\n{context}"
    )
    results = []
    for text, expected in SENTIMENT_TESTS:
        prompt = prompt_template.format(context=text)
        response, latency = call_model(model, prompt)
        score = score_exact(response, expected)
        results.append({
            "category": "Sentiment",
            "input": text[:80],
            "expected": expected,
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    return results


def run_extraction(model: str) -> list[dict]:
    results = []
    # Company name extraction
    for text, expected in EXTRACTION_COMPANY_TESTS:
        prompt = (
            "Extract the company name from this text. "
            "Reply with just the company name, nothing else.\n\nContext:\n" + text
        )
        response, latency = call_model(model, prompt)
        score = score_contains(response, expected)
        results.append({
            "category": "Extraction (Company)",
            "input": text[:80],
            "expected": expected,
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    # Dollar amount extraction
    for text, expected in EXTRACTION_AMOUNT_TESTS:
        prompt = (
            "Extract the dollar amount from this text as a plain number "
            "(expand abbreviations like K=thousand, M=million, B=billion). "
            "Reply with just the number, nothing else.\n\nContext:\n" + text
        )
        response, latency = call_model(model, prompt)
        score = score_amount(response, expected)
        results.append({
            "category": "Extraction (Amount)",
            "input": text[:80],
            "expected": expected,
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    return results


def run_categorization(model: str) -> list[dict]:
    prompt_template = (
        "Categorize this expense into exactly one of: Travel, Software, Meals, "
        "Marketing, Office. Reply with the category name only.\n\nContext:\n{context}"
    )
    results = []
    for text, expected in CATEGORIZATION_TESTS:
        prompt = prompt_template.format(context=text)
        response, latency = call_model(model, prompt)
        score = score_exact(response, expected)
        results.append({
            "category": "Categorization",
            "input": text[:80],
            "expected": expected,
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    return results


def run_cleaning(model: str) -> list[dict]:
    prompt_template = (
        "Clean and normalize this data entry. Reply with only the cleaned value, "
        "nothing else.\n\nContext:\n{context}"
    )
    results = []
    for messy, expected in CLEANING_TESTS:
        prompt = prompt_template.format(context=messy)
        response, latency = call_model(model, prompt)
        score = score_cleaning(response, expected)
        results.append({
            "category": "Data Cleaning",
            "input": messy,
            "expected": expected,
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    return results


def run_formulas(model: str) -> list[dict]:
    results = []
    for description, expected_funcs in FORMULA_TESTS:
        prompt = (
            "Write the Excel formula for this. Reply with just the formula, "
            "nothing else.\n\nContext:\n" + description
        )
        response, latency = call_model(model, prompt)
        score = score_formula(response, expected_funcs)
        results.append({
            "category": "Formula Helper",
            "input": description[:80],
            "expected": ", ".join(expected_funcs),
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    return results


def run_translation(model: str) -> list[dict]:
    results = []
    for english, keywords in TRANSLATION_TESTS:
        prompt = (
            "Translate this product description to Spanish. "
            "Reply with only the translation.\n\nContext:\n" + english
        )
        response, latency = call_model(model, prompt)
        score = score_keywords(response, keywords)
        results.append({
            "category": "Translation",
            "input": english[:80],
            "expected": f"Keywords: {', '.join(keywords[:3])}...",
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    return results


def run_code(model: str) -> list[dict]:
    results = []
    for description, language, keywords in CODE_TESTS:
        prompt = (
            f"Write a {language} code snippet that: {description}. "
            "Reply with only the code."
        )
        response, latency = call_model(model, prompt)
        score = score_code(response, language, keywords)
        results.append({
            "category": "Code Snippets",
            "input": f"[{language}] {description}",
            "expected": f"Keywords: {', '.join(keywords[:3])}...",
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    return results


def run_summarization(model: str) -> list[dict]:
    results = []
    for text, keywords in SUMMARIZATION_TESTS:
        prompt = (
            "Summarize this in one concise sentence.\n\nContext:\n" + text
        )
        response, latency = call_model(model, prompt)
        score = score_summary(response, keywords, len(text))
        results.append({
            "category": "Summarization",
            "input": text[:80] + "...",
            "expected": f"Keywords: {', '.join(keywords[:3])}...",
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    return results


def run_email(model: str) -> list[dict]:
    results = []
    for name, company, context in EMAIL_TESTS:
        prompt = (
            f"Write a short 3-sentence follow-up sales email.\n\n"
            f"Context:\nContact: {name}\nCompany: {company}\nNotes: {context}"
        )
        response, latency = call_model(model, prompt)
        score = score_email(response, name, company)
        results.append({
            "category": "Email Drafts",
            "input": f"{name} @ {company}",
            "expected": f"Mentions {name}, {company}",
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    return results


def run_product(model: str) -> list[dict]:
    results = []
    for product_name, specs, keywords in PRODUCT_TESTS:
        prompt = (
            "Write a 2-3 sentence marketing description for an e-commerce listing.\n\n"
            f"Context:\nProduct: {product_name}\nSpecs: {specs}"
        )
        response, latency = call_model(model, prompt)
        score = score_product(response, product_name, keywords)
        results.append({
            "category": "Product Copy",
            "input": product_name,
            "expected": f"Keywords: {', '.join(keywords[:3])}...",
            "response": response[:200],
            "score": score,
            "latency": latency,
        })
    return results


ALL_RUNNERS = [
    run_sentiment,
    run_extraction,
    run_categorization,
    run_cleaning,
    run_formulas,
    run_translation,
    run_code,
    run_summarization,
    run_email,
    run_product,
]


# ── Report Generation ──────────────────────────────────────────────────────

def generate_report(all_results: dict[str, list[dict]], output_path: str):
    """Generate Excel report with results."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Saving as JSON instead.")
        with open(output_path.replace(".xlsx", ".json"), "w") as f:
            json.dump(all_results, f, indent=2, default=str)
        return

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    score_font = Font(size=11)
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    mid_fill = PatternFill("solid", fgColor="FFEB9C")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")
    best_fill = PatternFill("solid", fgColor="92D050")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    categories = []
    for model_results in all_results.values():
        for r in model_results:
            if r["category"] not in categories:
                categories.append(r["category"])

    headers = ["Model"] + categories + ["Overall Score", "Avg Latency (ms)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Compute per-category scores for each model
    model_summaries = {}
    for row_idx, (model, results) in enumerate(all_results.items(), 2):
        cat_scores = {}
        cat_latencies = {}
        for r in results:
            cat = r["category"]
            cat_scores.setdefault(cat, []).append(r["score"])
            cat_latencies.setdefault(cat, []).append(r["latency"])

        avg_scores = {c: sum(s) / len(s) for c, s in cat_scores.items()}
        all_scores = [s for scores in cat_scores.values() for s in scores]
        all_latencies = [l for lats in cat_latencies.values() for l in lats]
        overall = sum(all_scores) / len(all_scores) if all_scores else 0
        avg_latency = sum(all_latencies) / len(all_latencies) if all_latencies else 0

        model_summaries[model] = {
            "scores": avg_scores,
            "overall": overall,
            "avg_latency": avg_latency,
        }

        ws.cell(row=row_idx, column=1, value=model).font = Font(bold=True, size=11)
        ws.cell(row=row_idx, column=1).border = thin_border

        for col_idx, cat in enumerate(categories, 2):
            score = avg_scores.get(cat, 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=round(score * 100, 1))
            cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="center")
            cell.font = score_font
            cell.border = thin_border
            if score >= 0.8:
                cell.fill = good_fill
            elif score >= 0.5:
                cell.fill = mid_fill
            else:
                cell.fill = bad_fill

        overall_col = len(categories) + 2
        latency_col = len(categories) + 3

        overall_cell = ws.cell(row=row_idx, column=overall_col, value=round(overall * 100, 1))
        overall_cell.number_format = "0.0"
        overall_cell.font = Font(bold=True, size=11)
        overall_cell.alignment = Alignment(horizontal="center")
        overall_cell.border = thin_border
        if overall >= 0.8:
            overall_cell.fill = good_fill
        elif overall >= 0.5:
            overall_cell.fill = mid_fill
        else:
            overall_cell.fill = bad_fill

        lat_cell = ws.cell(row=row_idx, column=latency_col, value=round(avg_latency * 1000))
        lat_cell.number_format = "#,##0"
        lat_cell.alignment = Alignment(horizontal="center")
        lat_cell.font = score_font
        lat_cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 20
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Highlight best overall score
    if model_summaries:
        best_model = max(model_summaries, key=lambda m: model_summaries[m]["overall"])
        for row_idx, model in enumerate(all_results.keys(), 2):
            if model == best_model:
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    if col == 1 or col == len(categories) + 2:
                        cell.fill = best_fill

    # ── Sheet 2: Detailed Results ──
    ws2 = wb.create_sheet("Detailed Results")
    detail_headers = ["Model", "Category", "Input", "Expected", "Response", "Score", "Latency (ms)"]
    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row = 2
    for model, results in all_results.items():
        for r in results:
            ws2.cell(row=row, column=1, value=model)
            ws2.cell(row=row, column=2, value=r["category"])
            ws2.cell(row=row, column=3, value=r["input"][:100])
            ws2.cell(row=row, column=4, value=str(r["expected"])[:100])
            ws2.cell(row=row, column=5, value=r["response"][:200])
            score_cell = ws2.cell(row=row, column=6, value=round(r["score"] * 100, 1))
            score_cell.number_format = "0.0"
            if r["score"] >= 0.8:
                score_cell.fill = good_fill
            elif r["score"] >= 0.5:
                score_cell.fill = mid_fill
            else:
                score_cell.fill = bad_fill
            ws2.cell(row=row, column=7, value=round(r["latency"] * 1000))
            row += 1

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 30
    ws2.column_dimensions["E"].width = 50
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 14

    # ── Sheet 3: Recommendation ──
    ws3 = wb.create_sheet("Recommendation")
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 50

    ws3.cell(row=1, column=1, value="Model Evaluation Report").font = Font(bold=True, size=16)
    ws3.merge_cells("A1:B1")

    if model_summaries:
        # Sort by overall score
        ranked = sorted(model_summaries.items(), key=lambda x: x[1]["overall"], reverse=True)

        ws3.cell(row=3, column=1, value="Overall Ranking").font = Font(bold=True, size=13)
        ws3.cell(row=4, column=1, value="Rank").font = Font(bold=True)
        ws3.cell(row=4, column=2, value="Model").font = Font(bold=True)
        ws3.cell(row=4, column=3, value="Score (%)").font = Font(bold=True)
        ws3.cell(row=4, column=4, value="Avg Latency (ms)").font = Font(bold=True)
        ws3.column_dimensions["C"].width = 14
        ws3.column_dimensions["D"].width = 18

        for i, (model, summary) in enumerate(ranked):
            r = 5 + i
            ws3.cell(row=r, column=1, value=f"#{i+1}")
            cell = ws3.cell(row=r, column=2, value=model)
            if i == 0:
                cell.font = Font(bold=True, color="006100")
                cell.fill = best_fill
            ws3.cell(row=r, column=3, value=round(summary["overall"] * 100, 1))
            ws3.cell(row=r, column=4, value=round(summary["avg_latency"] * 1000))

        # Recommendation
        rec_row = 5 + len(ranked) + 2
        best = ranked[0]
        ws3.cell(row=rec_row, column=1, value="Recommended Default").font = Font(bold=True, size=13)
        ws3.cell(row=rec_row + 1, column=1, value="Model:")
        ws3.cell(row=rec_row + 1, column=2, value=best[0]).font = Font(bold=True, size=12, color="006100")
        ws3.cell(row=rec_row + 2, column=1, value="Overall Score:")
        ws3.cell(row=rec_row + 2, column=2, value=f"{best[1]['overall']*100:.1f}%")
        ws3.cell(row=rec_row + 3, column=1, value="Avg Latency:")
        ws3.cell(row=rec_row + 3, column=2, value=f"{best[1]['avg_latency']*1000:.0f} ms")

        # Best per category
        cat_row = rec_row + 6
        ws3.cell(row=cat_row, column=1, value="Best Model per Category").font = Font(bold=True, size=13)
        cat_row += 1
        ws3.cell(row=cat_row, column=1, value="Category").font = Font(bold=True)
        ws3.cell(row=cat_row, column=2, value="Best Model").font = Font(bold=True)
        ws3.cell(row=cat_row, column=3, value="Score (%)").font = Font(bold=True)

        for cat in categories:
            cat_row += 1
            best_cat_model = max(
                model_summaries.items(),
                key=lambda x: x[1]["scores"].get(cat, 0),
            )
            ws3.cell(row=cat_row, column=1, value=cat)
            ws3.cell(row=cat_row, column=2, value=best_cat_model[0])
            ws3.cell(row=cat_row, column=3,
                     value=round(best_cat_model[1]["scores"].get(cat, 0) * 100, 1))

    wb.save(output_path)
    print(f"\nReport saved to: {output_path}")


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Excel AI model evaluation")
    parser.add_argument("--models", nargs="+", default=ALL_MODELS,
                        help="Models to test (default: all)")
    parser.add_argument("--cleanup", action="store_true",
                        help="Remove each model after testing to save disk space")
    parser.add_argument("--output", default="eval_results.xlsx",
                        help="Output Excel file path (default: eval_results.xlsx)")
    parser.add_argument("--skip-pull", action="store_true",
                        help="Skip pulling models (assume already downloaded)")
    args = parser.parse_args()

    print("=" * 70)
    print("  Excel AI — Comprehensive Model Evaluation")
    print("=" * 70)
    print(f"  Models: {len(args.models)}")
    print(f"  Test categories: {len(ALL_RUNNERS)}")
    total_tests = (len(SENTIMENT_TESTS) + len(EXTRACTION_COMPANY_TESTS) +
                   len(EXTRACTION_AMOUNT_TESTS) + len(CATEGORIZATION_TESTS) +
                   len(CLEANING_TESTS) + len(FORMULA_TESTS) +
                   len(TRANSLATION_TESTS) + len(CODE_TESTS) +
                   len(SUMMARIZATION_TESTS) + len(EMAIL_TESTS) +
                   len(PRODUCT_TESTS))
    print(f"  Tests per model: {total_tests}")
    print(f"  Total API calls: {total_tests * len(args.models)}")
    print(f"  Output: {args.output}")
    print("=" * 70)

    # Check Ollama is running
    try:
        urllib.request.urlopen(f"{OLLAMA_URL}/v1/models", timeout=5)
    except Exception:
        print("\nOllama is not running. Starting it...")
        subprocess.Popen(
            ["ollama", "serve"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        time.sleep(3)
        try:
            urllib.request.urlopen(f"{OLLAMA_URL}/v1/models", timeout=5)
        except Exception:
            print("ERROR: Cannot connect to Ollama. Is it installed?")
            sys.exit(1)

    all_results = {}
    failed_models = []

    for model_idx, model in enumerate(args.models, 1):
        print(f"\n{'─' * 70}")
        print(f"  [{model_idx}/{len(args.models)}] Testing: {model}")
        print(f"{'─' * 70}")

        # Pull model
        if not args.skip_pull:
            if not pull_model(model):
                failed_models.append(model)
                continue

        # Warm up (first call is often slower due to model loading)
        print("  Warming up...")
        call_model(model, "Say hello.")

        # Run all test categories
        model_results = []
        for runner in ALL_RUNNERS:
            cat_results = runner(model)
            cat_name = cat_results[0]["category"] if cat_results else "?"
            cat_scores = [r["score"] for r in cat_results]
            cat_avg = sum(cat_scores) / len(cat_scores) if cat_scores else 0
            cat_latencies = [r["latency"] for r in cat_results]
            cat_lat = sum(cat_latencies) / len(cat_latencies) if cat_latencies else 0
            print(f"  {cat_name:25s}  {cat_avg*100:5.1f}%  ({cat_lat*1000:.0f}ms avg)")
            model_results.extend(cat_results)

        all_results[model] = model_results

        # Overall for this model
        all_scores = [r["score"] for r in model_results]
        overall = sum(all_scores) / len(all_scores) if all_scores else 0
        all_latencies = [r["latency"] for r in model_results]
        avg_lat = sum(all_latencies) / len(all_latencies) if all_latencies else 0
        print(f"  {'OVERALL':25s}  {overall*100:5.1f}%  ({avg_lat*1000:.0f}ms avg)")

        # Cleanup if requested
        if args.cleanup:
            remove_model(model)

        # Save intermediate results
        try:
            generate_report(all_results, args.output)
        except Exception as e:
            print(f"  Warning: Could not save intermediate report: {e}")

    # Final report
    print(f"\n{'=' * 70}")
    print("  FINAL RESULTS")
    print(f"{'=' * 70}")

    if all_results:
        ranked = sorted(
            all_results.items(),
            key=lambda x: sum(r["score"] for r in x[1]) / len(x[1]),
            reverse=True,
        )
        for i, (model, results) in enumerate(ranked, 1):
            scores = [r["score"] for r in results]
            latencies = [r["latency"] for r in results]
            overall = sum(scores) / len(scores) * 100
            avg_lat = sum(latencies) / len(latencies) * 1000
            marker = " ← BEST" if i == 1 else ""
            print(f"  #{i:2d}  {model:22s}  {overall:5.1f}%  ({avg_lat:.0f}ms){marker}")

        generate_report(all_results, args.output)

    if failed_models:
        print(f"\n  Failed to pull: {', '.join(failed_models)}")

    print(f"\n  Report: {args.output}")
    print("  Done!")


if __name__ == "__main__":
    main()
